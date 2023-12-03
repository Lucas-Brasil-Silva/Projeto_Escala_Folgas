from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from calendar import monthcalendar
from datetime import datetime
from os import linesep

class Planilha:
    GREEN = '00008000'
    BLUE = '000000FF'
    CARAMEL = Font(color='A08708').color.rgb
    RED = Font(color='D92525').color.rgb

    def __init__(self,tabela,mes=None):
        """
        Inicializa a classe Planilha.

        Args:
            tabela (str): Nome do arquivo da planilha a ser gerada.
            mes (int, opcional): Mês da escala (1 a 12). O mês atual será usado se não especificado.

        A classe Planilha é usada para gerar e formatar uma planilha de escalas.

        """
        self.tabela = tabela
        self._configura_workbook(mes)
            
    def _configura_workbook(self, mes):
        """
        Configura a planilha.

        Args:
            mes (int, opcional): Mês da escala (1 a 12). O mês atual será usado se não especificado.

        Esta função configura a planilha, definindo o título, legendas e a estrutura básica da escala.

        """
        workbook = Workbook()
        ano = datetime.now().year
        mes =  datetime.now().month if not mes else mes
        workbook['Sheet'].title = 'Escala'
        sheet = workbook['Escala']
        sheet.merge_cells('A1:B1')
        sheet['A1'] = 'Legênda:'
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet['A1'].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        legendas = ['Folgas','HomeOffice','Plantão HomeOffice','Plantão Presencial']
        cores  = [self.GREEN,self.BLUE,self.CARAMEL,self.RED]
        for id,i in enumerate(range(2,6)):
            sheet[f'A{i}'].fill = PatternFill(start_color=cores[id], fill_type='solid')
            sheet[f'B{i}'] = legendas[id]
        for i in range(2,6):
            sheet[f'A{i}'].border = Border(left=Side(style='thin'),right=None,top=Side(style='thin'),bottom=Side(style='thin'))
            sheet[f'B{i}'].border = Border(left=None,right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        sheet.append([''])
        sheet.append([f'Data: {mes}/{ano}'])
        sheet.append(['Segunda-feira','Terça-feira','Quarta-feira','Quinta-feira','Sexta-feira','Sábado','Domingo'])
        matrix = monthcalendar(ano,mes)

        for dado in matrix:
            sheet.append(dado)
        workbook.save(self.tabela)

    def _colorir_texto(self,linha,cor,texto):
        """
        Formata o texto em uma célula da planilha com uma determinada cor.

        Args:
            linha (cell): A célula na qual o texto será formatado.
            cor (str): Código da cor a ser aplicada.
            texto (str): Texto a ser formatado na célula.

        """
        try:
            linha.value += CellRichText([linesep,TextBlock(cor,texto)])
        except TypeError:
            linha.value = CellRichText([f'{linha.value}',linesep,TextBlock(cor,texto)])

    def gerar_escala(self,folgas,homeoffice,plantoes):
        """
        Gera a escala na planilha, aplicando formatação de cores.

        Args:
            folgas (list): Lista de folgas.
            homeoffice (list): Lista de home office.
            plantoes (list): Lista de plantões.

        Esta função gera a escala na planilha e aplica formatação de cores para as células de acordo com as informações fornecidas nas listas de folgas, home office e plantões.

        """
        workbook = load_workbook(self.tabela)
        sheet = workbook[workbook.sheetnames[0]]
        
        for linhas in sheet.iter_rows(min_row=9):
            for linha in  linhas:
                folgas_ = [dado[0] for dado in folgas if linha.value == dado[1]]
                homeof = [dado[0] for dado in homeoffice if dado[1] == linha.value]
                p_plantao = [[plantao[0][0],plantao[1]] for plantao in plantoes if plantao[0][1]+2 == linha.value and plantao[1] == 'P-Plantão']
                h_plantao = [[plantao[0][0],plantao[1]] for plantao in plantoes if plantao[0][1] == linha.value and plantao[1] == 'H-Plantão']
                
                if folgas_:
                    self._colorir_texto(linha,InlineFont(color=self.GREEN),'\n'.join(folgas_))
                if homeof:
                    self._colorir_texto(linha,InlineFont(color=self.BLUE),'\n'.join(homeof))
                if p_plantao:
                    self._colorir_texto(linha,InlineFont(color=self.RED),' - '.join(p_plantao[0]))
                if h_plantao:
                    self._colorir_texto(linha,InlineFont(color=self.CARAMEL),' - '.join(h_plantao[0]))

        workbook.save(self.tabela)